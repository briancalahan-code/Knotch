#!/usr/bin/env python3
"""
Exec Report Data Engine — Knotch

Standalone script that calls HubSpot REST API, computes all report KPIs,
and outputs a deterministic JSON file. Claude prompts consume this JSON
to format the narrative report.

Usage:
  python3 compute_report.py --mode monthly --quarter Q2
  python3 compute_report.py --mode quarterly --quarter Q1 --dry-run
  python3 compute_report.py --mode monthly --quarter Q2 --output /tmp/my_report.json
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime, date, timezone
from zoneinfo import ZoneInfo

import requests

ET = ZoneInfo("America/New_York")

# --- Fiscal Calendar ---

FISCAL_QUARTERS = {
    "FY27": {
        "Q1": ("2026-02-01", "2026-04-30"),
        "Q2": ("2026-05-01", "2026-07-31"),
        "Q3": ("2026-08-01", "2026-10-31"),
        "Q4": ("2026-11-01", "2027-01-31"),
    },
    "FY26": {
        "Q1": ("2025-02-01", "2025-04-30"),
        "Q2": ("2025-05-01", "2025-07-31"),
        "Q3": ("2025-08-01", "2025-10-31"),
        "Q4": ("2025-11-01", "2026-01-31"),
    },
}

QUARTERLY_TARGETS = {
    "FY27": {
        "Q1": {"bookings": 1100000, "ipms": 24, "pipeline": 6250000},
        "Q2": {"bookings": 1100000, "ipms": 24, "pipeline": 6250000},
        "Q3": {"bookings": 1200000, "ipms": 24, "pipeline": 6250000},
        "Q4": {"bookings": 1200000, "ipms": 24, "pipeline": 6250000},
    }
}

# --- HubSpot Field Mapping ---

PIPELINE_ID = "72018330"

STAGES = {
    "152446547": "IPM",
    "152455272": "Qualification",
    "138620983": "Consensus",
    "138620984": "Proposal",
    "138620985": "Procurement",
    "138620988": "Closed Won",
    "138669962": "Closed Lost",
}

LATE_STAGES = ["138620984", "138620985"]
QUAL_PLUS = ["152455272", "138620983", "138620984", "138620985"]
OPEN_STAGES = ["152446547", "152455272", "138620983", "138620984", "138620985"]

OWNERS = {
    "693091902": "Don Vanderslice",
    "349190077": "Lee Fine",
    "81700088": "Tim Long",
    "87170480": "Pete Davies",
    "702586472": "Eli Grant",
    "723668113": "David Brown",
    "723668771": "Andrew Bolton",
    "627390764": "Ben Smith",
    "88616151": "Brian Calahan",
    "889074486": "Tommy Shaker",
    "758553440": "Ryan Ruxton",
    "2110079045": "Carolyn Scott",
}

NB_TEAM = ["693091902", "349190077", "81700088", "87170480"]


# --- HubSpot API Client ---


class HubSpotClient:
    BASE = "https://api.hubapi.com"

    def __init__(self, api_key, dry_run=False):
        self.headers = {
            "Authorization": f"Bearer {api_key}" if api_key else "",
            "Content-Type": "application/json",
        }
        self.dry_run = dry_run
        self._call_count = 0

    def search_deals(self, filters, properties, sorts=None, limit=200):
        return self._search("deals", filters, properties, sorts, limit)

    def search_objects(self, object_type, filters, properties, sorts=None, limit=200):
        return self._search(object_type, filters, properties, sorts, limit)

    def _search(self, object_type, filters, properties, sorts=None, limit=200):
        url = f"{self.BASE}/crm/v3/objects/{object_type}/search"
        body = {
            "filterGroups": [{"filters": filters}],
            "properties": properties,
            "limit": min(limit, 200),
        }
        if sorts:
            body["sorts"] = sorts

        if self.dry_run:
            print(f"[DRY-RUN] POST {url}")
            print(f"  Filters: {json.dumps(filters, indent=2)}")
            print(f"  Properties: {properties}")
            return []

        all_results = []
        after = None
        while True:
            if after:
                body["after"] = after
            self._call_count += 1
            resp = requests.post(url, headers=self.headers, json=body, timeout=30)

            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 10))
                print(f"  Rate limited, sleeping {retry_after}s...", file=sys.stderr)
                time.sleep(retry_after)
                continue

            if resp.status_code >= 400:
                print(
                    f"  ERROR {resp.status_code} on {url}: {resp.text[:300]}",
                    file=sys.stderr,
                )
            resp.raise_for_status()
            data = resp.json()
            results = data.get("results", [])
            all_results.extend(results)

            paging = data.get("paging", {}).get("next", {})
            after = paging.get("after")
            if not after:
                break

        return all_results


# --- Helper Functions ---


def parse_dollars(value):
    if not value:
        return 0
    try:
        return round(float(value))
    except (ValueError, TypeError):
        return 0


def to_et(iso_string):
    if not iso_string:
        return None
    dt = datetime.fromisoformat(iso_string)
    return dt.astimezone(ET)


def owner_name(owner_id):
    return OWNERS.get(str(owner_id), f"Unknown ({owner_id})")


def classify_deal_type(dealtype):
    if dealtype == "New License":
        return "New Biz"
    return "Upsell"


def classify_ace_k1(dealname):
    name_upper = (dealname or "").upper()
    if "ACE" in name_upper:
        return "ACE"
    return "K1"


def date_to_epoch_ms(date_str):
    dt = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=ET)
    return str(int(dt.astimezone(timezone.utc).timestamp() * 1000))


def end_of_day_epoch_ms(date_str):
    dt = datetime.strptime(date_str, "%Y-%m-%d").replace(
        hour=23, minute=59, second=59, tzinfo=ET
    )
    return str(int(dt.astimezone(timezone.utc).timestamp() * 1000))


def last_month_of_quarter(q_start, q_end):
    end_date = date.fromisoformat(q_end)
    start_of_last_month = end_date.replace(day=1)
    return str(start_of_last_month), q_end


# --- Configuration ---


def build_config(args):
    fy = args.fy
    q = args.quarter
    q_start, q_end = FISCAL_QUARTERS[fy][q]

    q_num = int(q[1])
    if q_num < 4:
        next_q = f"Q{q_num + 1}"
        next_q_fy = fy
    else:
        next_q = "Q1"
        next_q_fy = f"FY{int(fy[2:]) + 1}"
    next_q_start, next_q_end = FISCAL_QUARTERS.get(next_q_fy, {}).get(
        next_q, (None, None)
    )

    prior_fy = f"FY{int(fy[2:]) - 1}"
    prior_q_start, prior_q_end = FISCAL_QUARTERS.get(prior_fy, {}).get(q, (None, None))

    return {
        "mode": args.mode,
        "quarter": q,
        "fiscal_year": fy,
        "quarter_start": q_start,
        "quarter_end": q_end,
        "next_q": next_q,
        "next_q_start": next_q_start,
        "next_q_end": next_q_end,
        "prior_q_start": prior_q_start,
        "prior_q_end": prior_q_end,
        "targets": QUARTERLY_TARGETS.get(fy, {}).get(q, {}),
        "dry_run": args.dry_run,
        "ptm_path": os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "Brian - Copy of FY27 Leadership PTMs.xlsx",
        ),
    }


def build_meta(config):
    now = datetime.now(ET)
    return {
        "mode": config["mode"],
        "quarter": config["quarter"],
        "fiscal_year": config["fiscal_year"],
        "quarter_start": config["quarter_start"],
        "quarter_end": config["quarter_end"],
        "generated_at": now.isoformat(),
        "generated_at_display": now.strftime("%A, %B %d, %Y at %I:%M %p ET"),
    }


# --- Section Compute Functions (stubs — filled in Task 3 & 4) ---


def compute_bookings(client, config):
    # Q01: Closed-Won deals (current quarter)
    won_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {"propertyName": "dealstage", "operator": "EQ", "value": "138620988"},
            {
                "propertyName": "closedate",
                "operator": "GTE",
                "value": config["quarter_start"],
            },
            {
                "propertyName": "closedate",
                "operator": "LTE",
                "value": config["quarter_end"],
            },
        ],
        properties=[
            "dealname",
            "dealtype",
            "platform_amt",
            "closedate",
            "hubspot_owner_id",
            "manager_forecast",
            "manager_forecast_amount",
            "hs_manual_forecast_category",
        ],
    )

    # Q02: YoY comparison (same quarter last FY)
    yoy_total = 0
    if config["prior_q_start"]:
        prior_deals = client.search_deals(
            filters=[
                {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
                {"propertyName": "dealstage", "operator": "EQ", "value": "138620988"},
                {
                    "propertyName": "closedate",
                    "operator": "GTE",
                    "value": config["prior_q_start"],
                },
                {
                    "propertyName": "closedate",
                    "operator": "LTE",
                    "value": config["prior_q_end"],
                },
            ],
            properties=["platform_amt"],
        )
        yoy_total = sum(
            parse_dollars(d["properties"].get("platform_amt")) for d in prior_deals
        )

    # Q03: Open deals with manager forecast (current Q close dates)
    forecast_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "dealstage",
                "operator": "IN",
                "values": OPEN_STAGES,
            },
            {
                "propertyName": "closedate",
                "operator": "GTE",
                "value": config["quarter_start"],
            },
            {
                "propertyName": "closedate",
                "operator": "LTE",
                "value": config["quarter_end"],
            },
        ],
        properties=[
            "dealname",
            "dealtype",
            "platform_amt",
            "closedate",
            "hubspot_owner_id",
            "manager_forecast",
            "manager_forecast_amount",
            "hs_manual_forecast_category",
            "dealstage",
        ],
    )

    # Closed-won aggregations
    closed_won_total = sum(
        parse_dollars(d["properties"].get("platform_amt")) for d in won_deals
    )
    goal = config["targets"].get("bookings", 0)
    pct_to_goal = round(closed_won_total / goal * 100, 1) if goal else 0

    by_seller = {}
    by_type = {}
    deal_list = []
    for d in won_deals:
        p = d["properties"]
        amt = parse_dollars(p.get("platform_amt"))
        seller = owner_name(p.get("hubspot_owner_id"))
        dtype = classify_deal_type(p.get("dealtype", ""))
        by_seller[seller] = by_seller.get(seller, 0) + amt
        by_type[dtype] = by_type.get(dtype, 0) + amt
        deal_list.append(
            {
                "id": d["id"],
                "name": p.get("dealname", ""),
                "ae": seller,
                "type": dtype,
                "ace_k1": classify_ace_k1(p.get("dealname", "")),
                "close_date": p.get("closedate", ""),
                "arr": amt,
            }
        )

    # Manager forecast rollup
    forecast = {
        "commit": {"count": 0, "total": 0, "deals": []},
        "best_case": {"count": 0, "total": 0, "deals": []},
        "pipeline_cat": {"count": 0, "total": 0, "deals": []},
        "omit": {"count": 0, "total": 0},
        "missing": {"count": 0, "total": 0, "deals": []},
    }
    ic_forecast = {
        "commit": {"count": 0, "total": 0},
        "best_case": {"count": 0, "total": 0},
        "pipeline_cat": {"count": 0, "total": 0},
        "missing": {"count": 0, "total": 0},
    }

    for d in forecast_deals:
        p = d["properties"]
        deal_amt = parse_dollars(p.get("platform_amt"))
        mfa = (
            parse_dollars(p.get("manager_forecast_amount"))
            if p.get("manager_forecast_amount")
            else deal_amt
        )
        deal_info = {
            "name": p.get("dealname", ""),
            "ae": owner_name(p.get("hubspot_owner_id")),
            "arr": deal_amt,
            "stage": STAGES.get(p.get("dealstage", ""), ""),
            "close_date": p.get("closedate", ""),
            "forecast_cat": p.get("manager_forecast", ""),
        }

        # Manager forecast
        cat = (p.get("manager_forecast") or "").lower().strip()
        if cat == "commit":
            forecast["commit"]["count"] += 1
            forecast["commit"]["total"] += mfa
            forecast["commit"]["deals"].append(deal_info)
        elif cat in ("best case", "best_case"):
            forecast["best_case"]["count"] += 1
            forecast["best_case"]["total"] += mfa
            forecast["best_case"]["deals"].append(deal_info)
        elif cat == "pipeline":
            forecast["pipeline_cat"]["count"] += 1
            forecast["pipeline_cat"]["total"] += mfa
            forecast["pipeline_cat"]["deals"].append(deal_info)
        elif cat == "omit":
            forecast["omit"]["count"] += 1
            forecast["omit"]["total"] += mfa
        else:
            forecast["missing"]["count"] += 1
            forecast["missing"]["total"] += deal_amt
            forecast["missing"]["deals"].append(deal_info)

        # IC forecast
        ic_cat = (p.get("hs_manual_forecast_category") or "").lower().strip()
        if ic_cat == "commit":
            ic_forecast["commit"]["count"] += 1
            ic_forecast["commit"]["total"] += deal_amt
        elif ic_cat in ("best case", "best_case"):
            ic_forecast["best_case"]["count"] += 1
            ic_forecast["best_case"]["total"] += deal_amt
        elif ic_cat == "pipeline":
            ic_forecast["pipeline_cat"]["count"] += 1
            ic_forecast["pipeline_cat"]["total"] += deal_amt
        else:
            ic_forecast["missing"]["count"] += 1
            ic_forecast["missing"]["total"] += deal_amt

    forecast["total"] = (
        forecast["commit"]["total"]
        + forecast["best_case"]["total"]
        + forecast["pipeline_cat"]["total"]
    )
    ic_forecast["total"] = (
        ic_forecast["commit"]["total"]
        + ic_forecast["best_case"]["total"]
        + ic_forecast["pipeline_cat"]["total"]
    )

    yoy_pct = (
        round((closed_won_total - yoy_total) / yoy_total * 100, 1)
        if yoy_total
        else None
    )

    return {
        "closed_won_total": closed_won_total,
        "closed_won_count": len(won_deals),
        "quarterly_goal": goal,
        "pct_to_goal": pct_to_goal,
        "yoy_prior": yoy_total,
        "yoy_pct_change": yoy_pct,
        "by_seller": by_seller,
        "by_type": by_type,
        "deals": sorted(deal_list, key=lambda d: d["arr"], reverse=True),
        "forecast": forecast,
        "ic_forecast": ic_forecast,
    }


def compute_activity(client, config):
    q_start_ms = date_to_epoch_ms(config["quarter_start"])
    q_end_ms = end_of_day_epoch_ms(config["quarter_end"])

    # Q04: IPMs held (current quarter)
    ipm_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "ipm_held",
                "operator": "GTE",
                "value": q_start_ms,
            },
            {
                "propertyName": "ipm_held",
                "operator": "LTE",
                "value": q_end_ms,
            },
        ],
        properties=[
            "dealname",
            "dealtype",
            "platform_amt",
            "ipm_held",
            "hubspot_owner_id",
            "dealstage",
        ],
    )

    ipms_by_seller = {}
    ipm_deal_list = []
    for d in ipm_deals:
        p = d["properties"]
        seller = owner_name(p.get("hubspot_owner_id"))
        ipms_by_seller[seller] = ipms_by_seller.get(seller, 0) + 1
        ipm_deal_list.append(
            {
                "id": d["id"],
                "name": p.get("dealname", ""),
                "ae": seller,
                "type": classify_deal_type(p.get("dealtype", "")),
                "arr": parse_dollars(p.get("platform_amt")),
                "ipm_held": p.get("ipm_held", ""),
            }
        )

    # Q05: Pipeline created (deals entering Qualification in current Q)
    pipe_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "hs_v2_date_entered_152455272",
                "operator": "GTE",
                "value": q_start_ms,
            },
            {
                "propertyName": "hs_v2_date_entered_152455272",
                "operator": "LTE",
                "value": q_end_ms,
            },
        ],
        properties=[
            "dealname",
            "dealtype",
            "platform_amt",
            "hs_v2_date_entered_152455272",
            "hubspot_owner_id",
            "dealstage",
        ],
    )

    pipe_total = 0
    pipe_by_seller = {}
    pipe_by_type = {}
    pipe_deal_list = []
    for d in pipe_deals:
        p = d["properties"]
        amt = parse_dollars(p.get("platform_amt"))
        seller = owner_name(p.get("hubspot_owner_id"))
        ace_k1 = classify_ace_k1(p.get("dealname", ""))
        pipe_total += amt
        pipe_by_seller[seller] = pipe_by_seller.get(seller, 0) + amt
        pipe_by_type[ace_k1] = pipe_by_type.get(ace_k1, 0) + amt
        pipe_deal_list.append(
            {
                "id": d["id"],
                "name": p.get("dealname", ""),
                "ae": seller,
                "type": classify_deal_type(p.get("dealtype", "")),
                "ace_k1": ace_k1,
                "created_date": p.get("hs_v2_date_entered_152455272", ""),
                "arr": amt,
                "stage": STAGES.get(p.get("dealstage", ""), ""),
            }
        )

    # Q06: Sales emails (NB team, current quarter)
    email_results = client.search_objects(
        "emails",
        filters=[
            {"propertyName": "hs_timestamp", "operator": "GTE", "value": q_start_ms},
            {"propertyName": "hs_timestamp", "operator": "LTE", "value": q_end_ms},
            {"propertyName": "hubspot_owner_id", "operator": "IN", "values": NB_TEAM},
        ],
        properties=["hs_timestamp", "hubspot_owner_id"],
    )
    emails_by_seller = {}
    for e in email_results:
        seller = owner_name(e["properties"].get("hubspot_owner_id"))
        emails_by_seller[seller] = emails_by_seller.get(seller, 0) + 1

    # Q07: External meetings (NB team, current quarter)
    meeting_results = client.search_objects(
        "meetings",
        filters=[
            {"propertyName": "hs_timestamp", "operator": "GTE", "value": q_start_ms},
            {"propertyName": "hs_timestamp", "operator": "LTE", "value": q_end_ms},
            {"propertyName": "hubspot_owner_id", "operator": "IN", "values": NB_TEAM},
        ],
        properties=["hs_timestamp", "hubspot_owner_id"],
    )
    meetings_by_seller = {}
    for m in meeting_results:
        seller = owner_name(m["properties"].get("hubspot_owner_id"))
        meetings_by_seller[seller] = meetings_by_seller.get(seller, 0) + 1

    # Q11: Event Attendance (custom object, current quarter)
    try:
        event_results = client.search_objects(
            "2-62279031",
            filters=[
                {
                    "propertyName": "event_date",
                    "operator": "GTE",
                    "value": q_start_ms,
                },
                {
                    "propertyName": "event_date",
                    "operator": "LTE",
                    "value": q_end_ms,
                },
            ],
            properties=["event_name", "event_date", "event_status", "event_type"],
        )
    except Exception as e:
        print(f"  Event Attendance query failed (non-fatal): {e}", file=sys.stderr)
        event_results = []

    return {
        "ipms": {
            "total": len(ipm_deals),
            "goal": config["targets"].get("ipms", 24),
            "by_seller": ipms_by_seller,
            "deals": sorted(
                ipm_deal_list, key=lambda d: d.get("ipm_held", ""), reverse=True
            ),
        },
        "pipeline_created": {
            "total": pipe_total,
            "count": len(pipe_deals),
            "by_seller": pipe_by_seller,
            "by_type": pipe_by_type,
            "deals": sorted(pipe_deal_list, key=lambda d: d["arr"], reverse=True),
        },
        "emails": {
            "total": len(email_results),
            "by_seller": emails_by_seller,
        },
        "meetings": {
            "total": len(meeting_results),
            "by_seller": meetings_by_seller,
        },
        "events": {
            "count": len(event_results),
        },
    }


def compute_pipeline(client, config):
    # Q08: Late-stage deals — current Q close dates (Proposal+)
    late_stage_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "dealstage",
                "operator": "IN",
                "values": LATE_STAGES,
            },
            {
                "propertyName": "closedate",
                "operator": "GTE",
                "value": config["quarter_start"],
            },
            {
                "propertyName": "closedate",
                "operator": "LTE",
                "value": config["quarter_end"],
            },
        ],
        properties=[
            "dealname",
            "dealtype",
            "platform_amt",
            "closedate",
            "hubspot_owner_id",
            "manager_forecast",
            "manager_forecast_amount",
            "hs_manual_forecast_category",
            "dealstage",
        ],
    )

    late_total = sum(
        parse_dollars(d["properties"].get("platform_amt")) for d in late_stage_deals
    )
    late_deal_list = []
    for d in late_stage_deals:
        p = d["properties"]
        late_deal_list.append(
            {
                "id": d["id"],
                "name": p.get("dealname", ""),
                "ae": owner_name(p.get("hubspot_owner_id")),
                "type": classify_deal_type(p.get("dealtype", "")),
                "ace_k1": classify_ace_k1(p.get("dealname", "")),
                "stage": STAGES.get(p.get("dealstage", ""), ""),
                "close_date": p.get("closedate", ""),
                "arr": parse_dollars(p.get("platform_amt")),
                "manager_forecast": p.get("manager_forecast", ""),
            }
        )

    goal = config["targets"].get("bookings", 0)
    coverage = round(late_total / goal, 2) if goal else 0

    # Q09: Late-stage deals — next Q close dates
    next_q_late = []
    next_q_late_total = 0
    if config["next_q_start"]:
        next_late_deals = client.search_deals(
            filters=[
                {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
                {
                    "propertyName": "dealstage",
                    "operator": "IN",
                    "values": LATE_STAGES,
                },
                {
                    "propertyName": "closedate",
                    "operator": "GTE",
                    "value": config["next_q_start"],
                },
                {
                    "propertyName": "closedate",
                    "operator": "LTE",
                    "value": config["next_q_end"],
                },
            ],
            properties=[
                "dealname",
                "dealtype",
                "platform_amt",
                "closedate",
                "hubspot_owner_id",
                "manager_forecast",
                "dealstage",
            ],
        )
        next_q_late_total = sum(
            parse_dollars(d["properties"].get("platform_amt")) for d in next_late_deals
        )
        for d in next_late_deals:
            p = d["properties"]
            next_q_late.append(
                {
                    "id": d["id"],
                    "name": p.get("dealname", ""),
                    "ae": owner_name(p.get("hubspot_owner_id")),
                    "stage": STAGES.get(p.get("dealstage", ""), ""),
                    "close_date": p.get("closedate", ""),
                    "arr": parse_dollars(p.get("platform_amt")),
                }
            )

    # Q10: Early pipeline — next Q close dates (Qualification+)
    next_q_early_total = 0
    next_q_early_deals = []
    if config["next_q_start"]:
        early_deals = client.search_deals(
            filters=[
                {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
                {
                    "propertyName": "dealstage",
                    "operator": "IN",
                    "values": QUAL_PLUS,
                },
                {
                    "propertyName": "closedate",
                    "operator": "GTE",
                    "value": config["next_q_start"],
                },
                {
                    "propertyName": "closedate",
                    "operator": "LTE",
                    "value": config["next_q_end"],
                },
            ],
            properties=[
                "dealname",
                "dealtype",
                "platform_amt",
                "closedate",
                "hubspot_owner_id",
                "dealstage",
            ],
        )
        next_q_early_total = sum(
            parse_dollars(d["properties"].get("platform_amt")) for d in early_deals
        )
        for d in early_deals:
            p = d["properties"]
            next_q_early_deals.append(
                {
                    "id": d["id"],
                    "name": p.get("dealname", ""),
                    "ae": owner_name(p.get("hubspot_owner_id")),
                    "stage": STAGES.get(p.get("dealstage", ""), ""),
                    "close_date": p.get("closedate", ""),
                    "arr": parse_dollars(p.get("platform_amt")),
                }
            )

    return {
        "late_stage_current_q": {
            "total": late_total,
            "count": len(late_stage_deals),
            "deals": sorted(late_deal_list, key=lambda d: d["arr"], reverse=True),
        },
        "coverage_ratio": coverage,
        "next_q": {
            "late_stage_total": next_q_late_total,
            "late_stage_count": len(next_q_late),
            "early_total": next_q_early_total,
            "deals": sorted(
                next_q_late + next_q_early_deals,
                key=lambda d: d["arr"],
                reverse=True,
            ),
        },
    }


def compute_data_quality(client, config):
    # Query open late-stage deals for forecast coverage check
    late_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "dealstage",
                "operator": "IN",
                "values": LATE_STAGES,
            },
            {
                "propertyName": "closedate",
                "operator": "GTE",
                "value": config["quarter_start"],
            },
            {
                "propertyName": "closedate",
                "operator": "LTE",
                "value": config["quarter_end"],
            },
        ],
        properties=[
            "dealname",
            "platform_amt",
            "hubspot_owner_id",
            "dealstage",
            "manager_forecast",
            "closedate",
        ],
    )

    forecast_missing = []
    for d in late_deals:
        p = d["properties"]
        if not (p.get("manager_forecast") or "").strip():
            forecast_missing.append(
                {
                    "name": p.get("dealname", ""),
                    "ae": owner_name(p.get("hubspot_owner_id")),
                    "stage": STAGES.get(p.get("dealstage", ""), ""),
                    "arr": parse_dollars(p.get("platform_amt")),
                }
            )

    total_late = len(late_deals)
    with_forecast = total_late - len(forecast_missing)
    coverage_pct = round(with_forecast / total_late * 100, 1) if total_late else 100.0

    # Query Qual+ deals for null platform_amt and past-due checks
    qual_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "dealstage",
                "operator": "IN",
                "values": QUAL_PLUS,
            },
        ],
        properties=[
            "dealname",
            "platform_amt",
            "hubspot_owner_id",
            "dealstage",
            "closedate",
        ],
    )

    null_amt = []
    past_due = []
    today_str = date.today().isoformat()
    for d in qual_deals:
        p = d["properties"]
        if not (p.get("platform_amt") or "").strip():
            null_amt.append(
                {
                    "name": p.get("dealname", ""),
                    "ae": owner_name(p.get("hubspot_owner_id")),
                    "stage": STAGES.get(p.get("dealstage", ""), ""),
                }
            )
        cd = (p.get("closedate") or "")[:10]
        if cd and cd < today_str:
            past_due.append(
                {
                    "name": p.get("dealname", ""),
                    "ae": owner_name(p.get("hubspot_owner_id")),
                    "close_date": cd,
                    "stage": STAGES.get(p.get("dealstage", ""), ""),
                    "arr": parse_dollars(p.get("platform_amt")),
                }
            )

    return {
        "forecast_coverage_pct": coverage_pct,
        "forecast_warning": coverage_pct < 70,
        "forecast_missing_deals": forecast_missing,
        "null_platform_amt_deals": null_amt,
        "past_due_open_deals": past_due,
    }


def compute_market_feedback(client, config):
    # Q12: Closed-Lost deals (current Q, Qualification+ only)
    lost_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {"propertyName": "dealstage", "operator": "EQ", "value": "138669962"},
            {
                "propertyName": "closedate",
                "operator": "GTE",
                "value": config["quarter_start"],
            },
            {
                "propertyName": "closedate",
                "operator": "LTE",
                "value": config["quarter_end"],
            },
            {
                "propertyName": "hs_v2_date_entered_152455272",
                "operator": "HAS_PROPERTY",
            },
        ],
        properties=[
            "dealname",
            "dealtype",
            "platform_amt",
            "closedate",
            "hubspot_owner_id",
            "closed_lost_reason",
        ],
    )

    lost_total = 0
    lost_by_reason = {}
    lost_deal_list = []
    for d in lost_deals:
        p = d["properties"]
        amt = parse_dollars(p.get("platform_amt"))
        reason = p.get("closed_lost_reason", "") or "Not specified"
        lost_total += amt
        if reason not in lost_by_reason:
            lost_by_reason[reason] = {"count": 0, "total": 0}
        lost_by_reason[reason]["count"] += 1
        lost_by_reason[reason]["total"] += amt
        lost_deal_list.append(
            {
                "id": d["id"],
                "name": p.get("dealname", ""),
                "ae": owner_name(p.get("hubspot_owner_id")),
                "type": classify_deal_type(p.get("dealtype", "")),
                "ace_k1": classify_ace_k1(p.get("dealname", "")),
                "close_date": p.get("closedate", ""),
                "arr": amt,
                "reason": reason,
            }
        )

    # ACE/K1 mix — need won deals (requery Q01), lost deals (above), open Qual+ (Q13),
    # and created deals (use pipeline_created from activity if available, else requery)
    won_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {"propertyName": "dealstage", "operator": "EQ", "value": "138620988"},
            {
                "propertyName": "closedate",
                "operator": "GTE",
                "value": config["quarter_start"],
            },
            {
                "propertyName": "closedate",
                "operator": "LTE",
                "value": config["quarter_end"],
            },
        ],
        properties=["dealname", "dealtype", "platform_amt"],
    )

    q_start_ms = date_to_epoch_ms(config["quarter_start"])
    q_end_ms = end_of_day_epoch_ms(config["quarter_end"])
    created_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "hs_v2_date_entered_152455272",
                "operator": "GTE",
                "value": q_start_ms,
            },
            {
                "propertyName": "hs_v2_date_entered_152455272",
                "operator": "LTE",
                "value": q_end_ms,
            },
        ],
        properties=["dealname", "dealtype", "platform_amt"],
    )

    open_qual_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "dealstage",
                "operator": "IN",
                "values": QUAL_PLUS,
            },
        ],
        properties=["dealname", "dealtype", "platform_amt"],
    )

    def ace_k1_bucket(deals):
        ace = {"count": 0, "total": 0}
        k1 = {"count": 0, "total": 0}
        for d in deals:
            p = d["properties"]
            amt = parse_dollars(p.get("platform_amt"))
            if classify_ace_k1(p.get("dealname", "")) == "ACE":
                ace["count"] += 1
                ace["total"] += amt
            else:
                k1["count"] += 1
                k1["total"] += amt
        return {"ACE": ace, "K1": k1}

    return {
        "closed_lost": {
            "total": lost_total,
            "count": len(lost_deals),
            "deals": sorted(lost_deal_list, key=lambda d: d["arr"], reverse=True),
            "by_reason": lost_by_reason,
        },
        "ace_k1_mix": {
            "created": ace_k1_bucket(created_deals),
            "won": ace_k1_bucket(won_deals),
            "lost": ace_k1_bucket(lost_deals),
            "open": ace_k1_bucket(open_qual_deals),
        },
    }


def compute_seller_performance(client, config):
    q_start_ms = date_to_epoch_ms(config["quarter_start"])
    q_end_ms = end_of_day_epoch_ms(config["quarter_end"])

    # Closed-Won by seller
    won_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {"propertyName": "dealstage", "operator": "EQ", "value": "138620988"},
            {
                "propertyName": "closedate",
                "operator": "GTE",
                "value": config["quarter_start"],
            },
            {
                "propertyName": "closedate",
                "operator": "LTE",
                "value": config["quarter_end"],
            },
        ],
        properties=["platform_amt", "hubspot_owner_id"],
    )

    # Closed-Lost by seller (Qual+ only, for win rate)
    lost_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {"propertyName": "dealstage", "operator": "EQ", "value": "138669962"},
            {
                "propertyName": "closedate",
                "operator": "GTE",
                "value": config["quarter_start"],
            },
            {
                "propertyName": "closedate",
                "operator": "LTE",
                "value": config["quarter_end"],
            },
            {
                "propertyName": "hs_v2_date_entered_152455272",
                "operator": "HAS_PROPERTY",
            },
        ],
        properties=["platform_amt", "hubspot_owner_id"],
    )

    # Pipeline created by seller
    pipe_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "hs_v2_date_entered_152455272",
                "operator": "GTE",
                "value": q_start_ms,
            },
            {
                "propertyName": "hs_v2_date_entered_152455272",
                "operator": "LTE",
                "value": q_end_ms,
            },
        ],
        properties=["platform_amt", "hubspot_owner_id"],
    )

    # IPMs by seller
    ipm_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "ipm_held",
                "operator": "GTE",
                "value": q_start_ms,
            },
            {
                "propertyName": "ipm_held",
                "operator": "LTE",
                "value": q_end_ms,
            },
        ],
        properties=["hubspot_owner_id"],
    )

    # Q13: Open deals at Qual+ (for open pipeline and hygiene flags)
    open_deals = client.search_deals(
        filters=[
            {"propertyName": "pipeline", "operator": "EQ", "value": PIPELINE_ID},
            {
                "propertyName": "dealstage",
                "operator": "IN",
                "values": QUAL_PLUS,
            },
        ],
        properties=["platform_amt", "hubspot_owner_id", "hs_tag_ids"],
    )

    # Q14 + Q15: Emails and meetings in last month of quarter
    lm_start, lm_end = last_month_of_quarter(
        config["quarter_start"], config["quarter_end"]
    )
    lm_start_ms = date_to_epoch_ms(lm_start)
    lm_end_ms = end_of_day_epoch_ms(lm_end)

    lm_emails = client.search_objects(
        "emails",
        filters=[
            {"propertyName": "hs_timestamp", "operator": "GTE", "value": lm_start_ms},
            {"propertyName": "hs_timestamp", "operator": "LTE", "value": lm_end_ms},
            {"propertyName": "hubspot_owner_id", "operator": "IN", "values": NB_TEAM},
        ],
        properties=["hs_timestamp", "hubspot_owner_id"],
    )

    lm_meetings = client.search_objects(
        "meetings",
        filters=[
            {"propertyName": "hs_timestamp", "operator": "GTE", "value": lm_start_ms},
            {"propertyName": "hs_timestamp", "operator": "LTE", "value": lm_end_ms},
            {"propertyName": "hubspot_owner_id", "operator": "IN", "values": NB_TEAM},
        ],
        properties=["hs_timestamp", "hubspot_owner_id"],
    )

    # Build per-seller performance dict
    sellers = {}
    for oid in NB_TEAM:
        name = OWNERS[oid]
        sellers[name] = {
            "bookings": 0,
            "bookings_count": 0,
            "pipeline_created": 0,
            "pipeline_count": 0,
            "ipms": 0,
            "won_count": 0,
            "lost_count": 0,
            "win_rate": 0.0,
            "emails_last_month": 0,
            "meetings_last_month": 0,
            "open_pipeline": 0,
            "hygiene_flags": 0,
        }

    for d in won_deals:
        oid = str(d["properties"].get("hubspot_owner_id", ""))
        name = OWNERS.get(oid)
        if name and name in sellers:
            sellers[name]["bookings"] += parse_dollars(
                d["properties"].get("platform_amt")
            )
            sellers[name]["bookings_count"] += 1
            sellers[name]["won_count"] += 1

    for d in lost_deals:
        oid = str(d["properties"].get("hubspot_owner_id", ""))
        name = OWNERS.get(oid)
        if name and name in sellers:
            sellers[name]["lost_count"] += 1

    for d in pipe_deals:
        oid = str(d["properties"].get("hubspot_owner_id", ""))
        name = OWNERS.get(oid)
        if name and name in sellers:
            sellers[name]["pipeline_created"] += parse_dollars(
                d["properties"].get("platform_amt")
            )
            sellers[name]["pipeline_count"] += 1

    for d in ipm_deals:
        oid = str(d["properties"].get("hubspot_owner_id", ""))
        name = OWNERS.get(oid)
        if name and name in sellers:
            sellers[name]["ipms"] += 1

    for d in open_deals:
        oid = str(d["properties"].get("hubspot_owner_id", ""))
        name = OWNERS.get(oid)
        if name and name in sellers:
            sellers[name]["open_pipeline"] += parse_dollars(
                d["properties"].get("platform_amt")
            )
            if (d["properties"].get("hs_tag_ids") or "").strip():
                sellers[name]["hygiene_flags"] += 1

    for e in lm_emails:
        oid = str(e["properties"].get("hubspot_owner_id", ""))
        name = OWNERS.get(oid)
        if name and name in sellers:
            sellers[name]["emails_last_month"] += 1

    for m in lm_meetings:
        oid = str(m["properties"].get("hubspot_owner_id", ""))
        name = OWNERS.get(oid)
        if name and name in sellers:
            sellers[name]["meetings_last_month"] += 1

    # Compute win rates
    for name in sellers:
        s = sellers[name]
        total_decisions = s["won_count"] + s["lost_count"]
        s["win_rate"] = (
            round(s["won_count"] / total_decisions * 100, 1) if total_decisions else 0.0
        )
        del s["won_count"]
        del s["lost_count"]

    return sellers


def compute_initiatives(config):
    ptm_path = config["ptm_path"]
    if not os.path.exists(ptm_path):
        return {
            "ptm_file_date": None,
            "pete_ptms": [],
            "error": f"PTM file not found: {ptm_path}",
        }

    import openpyxl

    wb = openpyxl.load_workbook(ptm_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    file_mod = datetime.fromtimestamp(os.path.getmtime(ptm_path), tz=ET)

    pete_rows = []
    in_pete = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]
        cell_text = str(vals[0] or "").strip()
        if cell_text == "Pete":
            in_pete = True
            continue
        elif cell_text and in_pete:
            break
        if in_pete:
            pete_rows.append(
                {
                    "priority": str(vals[1] or "").strip(),
                    "tactics": str(vals[2] or "").strip(),
                    "metrics": str(vals[3] or "").strip(),
                    "due_date": str(vals[4] or "").strip(),
                    "status": str(vals[5] or "").strip(),
                    "latest_update": str(vals[6] or "").strip()
                    if len(vals) > 6
                    else "",
                }
            )

    return {
        "ptm_file_date": file_mod.strftime("%Y-%m-%d"),
        "ptm_sheet": wb.sheetnames[0],
        "pete_ptms": [r for r in pete_rows if r["priority"]],
    }


# --- Main ---


def parse_args():
    p = argparse.ArgumentParser(description="Knotch Exec Report Data Engine")
    p.add_argument(
        "--mode",
        required=True,
        choices=["monthly", "quarterly"],
        help="Report mode: monthly (sections 1-3) or quarterly (sections 1-6)",
    )
    p.add_argument(
        "--quarter",
        required=True,
        choices=["Q1", "Q2", "Q3", "Q4"],
        help="Fiscal quarter to report on",
    )
    p.add_argument("--fy", default="FY27", help="Fiscal year (default: FY27)")
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Print API calls without executing",
    )
    p.add_argument(
        "--output",
        default="/tmp/exec_report_data.json",
        help="Output JSON path (default: /tmp/exec_report_data.json)",
    )
    return p.parse_args()


def main():
    args = parse_args()
    api_key = os.environ.get("HUBSPOT_API_KEY")
    if not api_key and not args.dry_run:
        print("ERROR: HUBSPOT_API_KEY environment variable not set", file=sys.stderr)
        sys.exit(1)

    config = build_config(args)
    client = HubSpotClient(api_key, dry_run=args.dry_run)

    report = {
        "meta": build_meta(config),
        "targets": config["targets"],
        "data_quality": {},
    }

    report["bookings"] = compute_bookings(client, config)
    report["activity"] = compute_activity(client, config)
    report["pipeline"] = compute_pipeline(client, config)
    report["data_quality"] = compute_data_quality(client, config)

    if args.mode == "quarterly":
        report["market_feedback"] = compute_market_feedback(client, config)
        report["seller_performance"] = compute_seller_performance(client, config)
        report["initiatives"] = compute_initiatives(config)

    with open(args.output, "w") as f:
        json.dump(report, f, indent=2, default=str)
    print(f"Report written to {args.output}")


if __name__ == "__main__":
    main()
